"""Inspect template.xlsx: formulas, values, merged cells, named ranges."""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\nikhi\Desktop\AuditReport\template.xlsx"

# Load with data_only=False to see formulas
wb = load_workbook(PATH, data_only=False)

print("=" * 80)
print("SHEETS IN WORKBOOK:", wb.sheetnames)
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name!r}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"Merged cells: {list(ws.merged_cells.ranges)}")
    print('=' * 80)

    # --- Part 1: All formula cells (any row) ---
    print(f"\n--- [ALL FORMULA CELLS IN {sheet_name!r}] ---")
    formula_count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                print(f"  {cell.coordinate}: {cell.value}")
                formula_count += 1
    if formula_count == 0:
        print("  (no formulas found anywhere)")

    # --- Part 2: All non-None cells in rows 70-100 ---
    print(f"\n--- [NON-NONE CELLS IN ROWS 70-100 IN {sheet_name!r}] ---")
    count_70_100 = 0
    for row_idx in range(70, min(101, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                col_letter = get_column_letter(col_idx)
                val_repr = repr(cell.value)
                print(f"  {col_letter}{row_idx}: {val_repr}")
                count_70_100 += 1
    if count_70_100 == 0:
        print("  (all cells empty in rows 70-100)")

print(f"\n{'=' * 80}")
print("NAMED RANGES / DEFINED NAMES")
print('=' * 80)
if wb.defined_names:
    for dn in wb.defined_names.definedName:
        print(f"  Name: {dn.name}, Attr: {dn.attr_text}")
else:
    print("  (none)")

# --- Also check with data_only=True for cached values ---
print(f"\n{'=' * 80}")
print("CACHED VALUES (data_only=True) - ROWS 70-85 FOR ALL SHEETS")
print('=' * 80)
wb2 = load_workbook(PATH, data_only=True)
for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    print(f"\n--- [{sheet_name}] ---")
    for row_idx in range(70, min(86, ws2.max_row + 1)):
        for col_idx in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                col_letter = get_column_letter(col_idx)
                print(f"  {col_letter}{row_idx}: {cell.value}")

wb.close()
wb2.close()
print("\nDone.")
