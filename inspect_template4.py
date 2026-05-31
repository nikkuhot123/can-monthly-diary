"""Check which rows actually have data in columns L, N, O, P, Q to validate formula ranges."""

from openpyxl import load_workbook

PATH = r"C:\Users\nikhi\Desktop\AuditReport\template.xlsx"
wb = load_workbook(PATH, data_only=False)
ws = wb["Sheet2"]

col_map = {
    'J': 10,  # Fare
    'K': 11,  # Lodging
    'L': 12,  # Boarding
    'M': 13,  # H.A.
    'N': 14,  # Diem Allowance
    'O': 15,  # Mode of Conveyance
    'P': 16,  # Distance in KMs
    'Q': 17,  # Fare
    'R': 18,  # Duty Timings
    'S': 19,  # Depature Time / Grand Total?
}

print("=" * 120)
print("SHEET2: Full data audit for cols J-S across rows 13-77")
print("=" * 120)
print(f"{'Row':<6}", end="")
for col_name in ['J','K','L','M','N','O','P','Q','R','S']:
    print(f"{col_name:>12}", end="")
print()

for row_idx in range(13, 78):
    # Check if row has ANY content
    row_has_financial = False
    for col_idx in range(10, 20):  # J through S
        if ws.cell(row=row_idx, column=col_idx).value is not None:
            row_has_financial = True
            break
    if not row_has_financial:
        continue

    print(f"Row {row_idx:<3}", end="")
    for col_name, col_idx in col_map.items():
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            print(f"{repr(cell.value):>12}", end="")
        else:
            print(f"{'':>12}", end="")
    print()

print()
print("=" * 120)
print("SHEET2: Also check if columns L, N, O, P, Q have ANY data ANYWHERE in rows 1-84")
print("=" * 120)
for col_name, col_idx in col_map.items():
    if col_name in ['L','N','O','P','Q']:
        found_rows = []
        for row_idx in range(1, 85):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                found_rows.append(row_idx)
        if found_rows:
            print(f"  Column {col_name}: data in rows {found_rows}")
        else:
            print(f"  Column {col_name}: NO DATA ANYWHERE in the sheet")

print()
print("=" * 120)
print("SHEET2: Check what S78 formula actually evaluates to vs intent")
print("=" * 120)
# S78 = SUM(J78:R78) — this sums J78 through R78 (columns 10-18)
# Let's list all totals in row 78
print("Row 78 values:")
for col_idx in range(10, 20):
    col_letter = chr(64 + col_idx)  # approximate
    # Better way
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col_idx)
    cell = ws.cell(row=78, column=col_idx)
    print(f"  {col_letter}78: {repr(cell.value)}")

print()
# Check what R column actually contains in data rows
print("=" * 120)
print("SHEET2: R column (Duty Timings) data rows — any numeric values?")
print("=" * 120)
for row_idx in range(13, 78):
    r_cell = ws.cell(row=row_idx, column=18)
    if r_cell.value is not None:
        print(f"  R{row_idx}: {repr(r_cell.value)}")

print()
print("=" * 120)
print("SHEET2: S column (Depature Time / Remarks) data rows — any values?")
print("=" * 120)
for row_idx in range(13, 78):
    s_cell = ws.cell(row=row_idx, column=19)
    if s_cell.value is not None:
        print(f"  S{row_idx}: {repr(s_cell.value)}")

wb.close()
print("\nDone.")
