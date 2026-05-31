import openpyxl
import os
import calendar as py_cal
from datetime import datetime, date, time, timedelta
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Alignment
from services.calendar_utils import get_bank_holidays_for_month

TEMPLATE_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "template.xlsx")
)


def _resolve_template_path():
    if os.path.exists(TEMPLATE_PATH):
        return TEMPLATE_PATH
    raise FileNotFoundError(f"Diary Excel template not found: {TEMPLATE_PATH}")


def _set_cell_value(ws, row_or_cell, column=None, value=None):
    if column is None:
        cell_ref = row_or_cell
        cell = ws[cell_ref]
    else:
        cell = ws.cell(row=row_or_cell, column=column)
        cell_ref = cell.coordinate

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).value = value
                return

    cell.value = value


def _get_attr(obj, *names, default=None):
    for name in names:
        if hasattr(obj, name):
            value = getattr(obj, name)
            if value not in (None, ""):
                return value
    return default


def _safe_upper(value, default=""):
    return str(value or default).upper()


def _as_time(value):
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str) and value.strip():
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                pass
    return value or None


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    return None


def _money(*values):
    return sum(float(value or 0) for value in values)


def _date_list_text(dates):
    dates = sorted(d for d in dates if d)
    if not dates:
        return ""
    return ",".join(str(d.day) for d in dates) + ","


def _clear_cells(ws, rows, columns):
    for row in rows:
        for column in columns:
            _set_cell_value(ws, row, column, value=None)


def _autofit_column_by_content(ws, column_letter, values, min_width=11, max_width=28, padding=2):
    content_lengths = [len(str(v).strip()) for v in values if str(v or "").strip()]
    if not content_lengths:
        ws.column_dimensions[column_letter].width = min_width
        return
    target_width = max(content_lengths) + padding
    ws.column_dimensions[column_letter].width = max(min_width, min(max_width, target_width))


def _configure_a4_print(ws):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def _build_duty_groups(attendance, diary):
    holidays = set(get_bank_holidays_for_month(diary.year, diary.month).keys())
    records = []
    for record in attendance:
        duty_date = _as_date(record.duty_date)
        if not duty_date or record.is_leave or record.is_holiday:
            continue
        records.append(record)

    records.sort(key=lambda record: _as_date(record.duty_date))
    groups = []
    for record in records:
        duty_date = _as_date(record.duty_date)
        key = (
            (record.branch_name or "").strip().upper(),
            (record.dp_code or "").strip(),
            (record.audit_type or "").strip().upper(),
            (record.place or "").strip().upper(),
        )

        if groups:
            previous = groups[-1]
            previous_date = _as_date(previous["records"][-1].duty_date)
            gap_dates = [
                previous_date + timedelta(days=offset)
                for offset in range(1, (duty_date - previous_date).days)
            ]
            gap_is_only_holidays = all(d in holidays or d.weekday() == 6 for d in gap_dates)
            if previous["key"] == key and (duty_date == previous_date + timedelta(days=1) or gap_is_only_holidays):
                previous["records"].append(record)
                previous["dates"].append(duty_date)
                continue

        groups.append({"key": key, "records": [record], "dates": [duty_date]})
    return groups


def _write_sheet1(ws, user, diary, attendance):
    user_dp = user.dp_code or ""
    today = date.today()
    month_label = f"{py_cal.month_abbr[diary.month].upper()}-{diary.year}"
    camp = _safe_upper(user.section, "NASHIK")
    name = _safe_upper(user.name)

    _set_cell_value(ws, "C7", value=name)
    _set_cell_value(ws, "C8", value=user.staff_no)
    _set_cell_value(ws, "C10", value=_safe_upper(user.zone, "MUMBAI"))
    _set_cell_value(ws, "C11", value=_safe_upper(user.section, "NASHIK"))
    _set_cell_value(ws, "C12", value=camp)
    # Personal details (address, bank, TA ID)
    _set_cell_value(ws, "P8", value=user.address_line1)
    _set_cell_value(ws, "P9", value=user.address_line2)
    _set_cell_value(ws, "P10", value=user.city_pin)
    _set_cell_value(ws, "C9", value=user.bank_name)
    _set_cell_value(ws, "T17", value=user.ta_id)
    _set_cell_value(ws, "P13", value=f"DATE : {today.strftime('%d.%m.%Y')}")
    _set_cell_value(ws, "P21", value=month_label)
    _set_cell_value(ws, "U21", value="=Sheet3!R24")
    _set_cell_value(ws, "C128", value=camp)
    _set_cell_value(ws, "C129", value=today.strftime("%d.%m.%Y"))
    _set_cell_value(ws, "T131", value=name)

    duty_groups = _build_duty_groups(attendance, diary)
    _clear_cells(ws, range(29, 43), [1, 2, 8, 10, 15, 18, 21])
    _clear_cells(ws, range(50, 64), [1, 2, 8, 10, 12, 15, 18, 21])
    _clear_cells(ws, range(68, 81), [1, 2, 8, 10, 13, 16, 20])

    for idx, group in enumerate(duty_groups[:14], start=1):
        row = 28 + idx
        first_record = group["records"][0]
        start_date = min(group["dates"])
        end_date = max(group["dates"])
        _set_cell_value(ws, row, 1, value=idx)
        _set_cell_value(ws, row, 2, value=_safe_upper(first_record.branch_name, "OFFICE"))
        _set_cell_value(ws, row, 8, value=first_record.dp_code or user_dp)
        _set_cell_value(ws, row, 10, value=_safe_upper(first_record.audit_type, "DUTY"))
        _set_cell_value(ws, row, 15, value=start_date)
        _set_cell_value(ws, row, 18, value=end_date)
        _set_cell_value(ws, row, 21, value=f"=B{row}")

        timing_row = 49 + idx
        _set_cell_value(ws, timing_row, 1, value=idx)
        _set_cell_value(ws, timing_row, 2, value=_safe_upper(first_record.branch_name, "OFFICE"))
        _set_cell_value(ws, timing_row, 8, value=first_record.dp_code or user_dp)
        _set_cell_value(ws, timing_row, 10, value=_as_time(first_record.duty_from_time))
        _set_cell_value(ws, timing_row, 12, value=_as_time(first_record.duty_to_time))

        manday_row = 67 + idx
        _set_cell_value(ws, manday_row, 1, value=idx)
        _set_cell_value(ws, manday_row, 2, value=_safe_upper(first_record.branch_name, "OFFICE"))
        _set_cell_value(ws, manday_row, 8, value=first_record.dp_code or user_dp)
        _set_cell_value(ws, manday_row, 10, value=start_date)
        _set_cell_value(ws, manday_row, 13, value=end_date)
        _set_cell_value(ws, manday_row, 16, value=len(group["dates"]))
        _set_cell_value(ws, manday_row, 20, value=_date_list_text(group["dates"]))

    _set_cell_value(ws, "P81", value="=SUM(P68:S80)")

    leave_dates = sorted(
        _as_date(record.duty_date)
        for record in attendance
        if record.is_leave and _as_date(record.duty_date)
    )
    _clear_cells(ws, range(87, 93), [10, 13, 16, 18, 20])
    for row in range(87, 93):
        _set_cell_value(ws, row, 16, value=0)
        _set_cell_value(ws, row, 18, value=0)
    _set_cell_value(ws, "P88", value=len(leave_dates))
    _set_cell_value(ws, "R88", value=len(leave_dates))
    _set_cell_value(ws, "T88", value=_date_list_text(leave_dates))
    _set_cell_value(ws, "P93", value="=SUM(P87:Q92)")
    _set_cell_value(ws, "R93", value="=SUM(R87:S92)")

    holidays = get_bank_holidays_for_month(diary.year, diary.month)
    extra_holidays = {
        _as_date(record.duty_date): "Public Holiday"
        for record in attendance
        if record.is_holiday and _as_date(record.duty_date)
    }
    holidays.update(extra_holidays)
    sundays = [d for d in holidays if d.weekday() == 6]
    saturdays = [d for d, reason in holidays.items() if "Saturday" in reason]
    public_holidays = [
        d for d, reason in holidays.items()
        if d.weekday() != 6 and "Saturday" not in reason
    ]

    for row in range(98, 102):
        _set_cell_value(ws, row, 13, value=None)
        _set_cell_value(ws, row, 16, value=None)
        _set_cell_value(ws, row, 20, value=None)
    for row, dates in ((98, public_holidays), (99, sundays), (100, saturdays)):
        _set_cell_value(ws, row, 13, value=len(dates))
        _set_cell_value(ws, row, 16, value=len(dates))
        _set_cell_value(ws, row, 20, value=_date_list_text(dates))
    _set_cell_value(ws, "M102", value="=SUM(M98:O101)")
    _set_cell_value(ws, "P102", value="=SUM(P98:S101)")

    _set_cell_value(ws, "P107", value="=P81")
    _set_cell_value(ws, "T107", value="=CONCATENATE(T68,T69,T70,T71,T72,T73,T74,T75,T76,T77,T78,T79,T80,T81)")
    _set_cell_value(ws, "P108", value="=P93")
    _set_cell_value(ws, "T108", value="=T88")
    _set_cell_value(ws, "P109", value="=P102")
    _set_cell_value(ws, "T109", value="=CONCATENATE(T98,T99,T100)")
    _set_cell_value(ws, "P110", value="=SUM(P107:S109)")


def _write_sheet2(ws, user, diary, attendance, travel, hotels, local):
    data_start_row = 13
    data_end_row = 77
    data_block_height = 2
    today = date.today()
    month_label = f"{py_cal.month_abbr[diary.month].upper()} {diary.year}"
    camp = _safe_upper(user.section, "NASHIK")
    user_dp = user.dp_code or ""

    _set_cell_value(ws, "K4", value=month_label)
    _set_cell_value(ws, "T4", value=today.strftime("%d.%m.%Y"))
    _set_cell_value(ws, "C6", value="=Sheet1!C7")
    _set_cell_value(ws, "G6", value="=Sheet1!C8")
    _set_cell_value(ws, "I6", value=f"DESIGNATION:  {_safe_upper(user.designation)}")
    _set_cell_value(ws, "T6", value=camp)
    _set_cell_value(ws, "G8", value=user.basic_pay or 0)
    _set_cell_value(ws, "A82", value=f"PLACE: {camp}")
    _set_cell_value(ws, "A83", value=f"DATE: {today.strftime('%d.%m.%Y')}")

    travel_map = {}
    for t in travel:
        d = _as_date(t.date_start)
        if d:
            travel_map.setdefault(d, []).append(t)

    hotel_map = {}
    for h in hotels:
        d = _as_date(h.checkin_date)
        if not d:
            continue
        totals = hotel_map.setdefault(d, {"lodging": 0, "boarding": 0, "ha": 0})
        totals["lodging"] += float(h.lodging_amount or 0)
        totals["boarding"] += float(h.boarding_amount or 0)
        totals["ha"] += float(h.halting_allowance or 0)

    local_map = {}
    for item in local:
        d = _as_date(item.travel_date)
        if not d:
            continue
        totals = local_map.setdefault(d, {"modes": [], "distance": 0, "fare": 0})
        if item.mode:
            totals["modes"].append(item.mode)
        totals["distance"] += float(item.distance_km or 0)
        totals["fare"] += _money(item.claimed_amount_exc_gst, item.claimed_gst)

    for r_idx in range(data_start_row, data_end_row + 1):
        for c_idx in range(1, 18):
            _set_cell_value(ws, r_idx, c_idx, value=None)

    attendance_map = {
        _as_date(record.duty_date): record
        for record in attendance
        if _as_date(record.duty_date)
    }
    holidays = get_bank_holidays_for_month(diary.year, diary.month)
    _, days_in_month = py_cal.monthrange(diary.year, diary.month)

    running_totals = {"fare": 0, "lodging": 0, "boarding": 0, "ha": 0, "diem": 0, "local_dist": 0, "local_fare": 0}
    purpose_values = []

    current_row = data_start_row
    for day_num in range(1, days_in_month + 1):
        if current_row > data_end_row:
            break

        d = date(diary.year, diary.month, day_num)
        record = attendance_map.get(d)
        purpose = ""
        if record:
            purpose = record.branch_name or "OFFICE"
            if record.is_holiday:
                purpose = holidays.get(d, "BANK HOLIDAY")
            if record.is_leave:
                purpose = "LEAVE"
        elif d in holidays:
            purpose = holidays[d].upper()
        else:
            purpose = "MISSING"

        h_rec = hotel_map.get(d, {})
        lc_rec = local_map.get(d, {})

        row_fare = sum(t.total_amount or 0 for t in travel_map.get(d, []))
        running_totals["fare"] += row_fare
        running_totals["lodging"] += h_rec.get("lodging", 0)
        running_totals["boarding"] += h_rec.get("boarding", 0)
        running_totals["ha"] += h_rec.get("ha", 0)
        running_totals["local_fare"] += lc_rec.get("fare", 0)

        _set_cell_value(ws, current_row, 1, value=day_num)
        _set_cell_value(ws, current_row, 2, value=_safe_upper(purpose))
        purpose_values.append(_safe_upper(purpose))
        purpose_cell = ws.cell(row=current_row, column=2)
        current_align = purpose_cell.alignment or Alignment()
        purpose_cell.alignment = Alignment(
            horizontal=current_align.horizontal,
            vertical="center",
            text_rotation=current_align.text_rotation,
            wrap_text=False,
            shrink_to_fit=True,
            indent=current_align.indent,
        )
        if purpose and len(str(purpose)) > 20:
            existing_height = ws.row_dimensions[current_row].height or 15
            ws.row_dimensions[current_row].height = max(existing_height, 20)
        _set_cell_value(ws, current_row, 11, value=h_rec.get("lodging") or None)
        _set_cell_value(ws, current_row, 12, value=h_rec.get("boarding") or None)
        _set_cell_value(ws, current_row, 13, value=h_rec.get("ha") or None)
        _set_cell_value(ws, current_row, 15, value=", ".join(lc_rec.get("modes", [])) or None)
        _set_cell_value(ws, current_row, 16, value=lc_rec.get("distance") or None)
        _set_cell_value(ws, current_row, 17, value=lc_rec.get("fare") or None)

        for i, t in enumerate(travel_map.get(d, [])[:data_block_height]):
            row_idx = current_row + i
            mode = " ".join(part for part in [t.mode, t.travel_class] if part)
            _set_cell_value(ws, row_idx, 3, value="Onward" if i == 0 else "Return")
            _set_cell_value(ws, row_idx, 4, value=t.from_place)
            _set_cell_value(ws, row_idx, 5, value=_as_time(_get_attr(t, "departure_time", "time_start")))
            _set_cell_value(ws, row_idx, 6, value=t.to_place)
            _set_cell_value(ws, row_idx, 7, value=_as_time(_get_attr(t, "arrival_time", "time_arrival")))
            _set_cell_value(ws, row_idx, 8, value=t.distance_km)
            _set_cell_value(ws, row_idx, 9, value=mode)
            _set_cell_value(ws, row_idx, 10, value=t.total_amount)

        current_row += data_block_height

    # Fix template SUM formulas at row 78 to use correct ranges (13:77 instead of 31:77)
    # Template has e.g. =SUM(L31:L77) — replace "31:" with "13:" to handle all columns
    for col_letter in ["L", "N", "O", "P", "Q"]:
        cell = ws[f"{col_letter}78"]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("=SUM("):
            old_range = f"{col_letter}31:{col_letter}77"
            new_range = f"{col_letter}13:{col_letter}77"
            cell.value = cell.value.replace(old_range, new_range)

    _autofit_column_by_content(ws, "B", purpose_values, min_width=11, max_width=28, padding=2)


def _write_sheet3(ws, user, other):
    today = date.today()
    camp = _safe_upper(user.section, "NASHIK")
    name = _safe_upper(user.name)

    _set_cell_value(ws, "C1", value=f"{name} ({user.staff_no})")
    _set_cell_value(ws, "C2", value=_safe_upper(user.zone, "MUMBAI"))
    _set_cell_value(ws, "C3", value=camp)
    _set_cell_value(ws, "O1", value=user.bank_name)
    _set_cell_value(ws, "B35", value=user.bank_account_no)
    _set_cell_value(ws, "B36", value=name)
    _set_cell_value(ws, "C41", value=camp)
    _set_cell_value(ws, "C42", value=today)
    _set_cell_value(ws, "E42", value=name)

    for row in range(7, 14):
        for column in (2, 3, 7):
            _set_cell_value(ws, row, column, value=None)

    for i, exp in enumerate(other[:7]):
        row_idx = 7 + i
        expense_date = _as_date(_get_attr(exp, "expense_date", "invoice_date", "created_at"))
        amount = _money(exp.claimed_amount_exc_gst, exp.claimed_gst, exp.declaration_amount)
        _set_cell_value(ws, row_idx, 2, value=expense_date)
        _set_cell_value(ws, row_idx, 3, value=exp.expense_description)
        _set_cell_value(ws, row_idx, 7, value=amount or None)

    _set_cell_value(ws, "G14", value="=SUM(G7:G13)")
    _set_cell_value(ws, "R7", value="=Sheet2!J78")
    _set_cell_value(ws, "R9", value="=Sheet2!K78")
    _set_cell_value(ws, "R11", value="=Sheet2!L78")
    _set_cell_value(ws, "R13", value="=Sheet2!M78")
    _set_cell_value(ws, "R15", value="=Sheet2!N78")
    _set_cell_value(ws, "R17", value="=Sheet2!Q78")
    _set_cell_value(ws, "R19", value="=G14")
    _set_cell_value(ws, "R20", value="=SUM(R7:R19)")
    _set_cell_value(ws, "R22", value=0)
    _set_cell_value(ws, "R24", value="=R20-R22")


def generate_diary_excel(fpath, user, diary, attendance, travel, hotels, local, other):
    template_path = _resolve_template_path()

    wb = openpyxl.load_workbook(template_path)
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    
    if 'Sheet1' in wb.sheetnames:
        _write_sheet1(wb['Sheet1'], user, diary, attendance)

    if 'Sheet2' in wb.sheetnames:
        _write_sheet2(wb['Sheet2'], user, diary, attendance, travel, hotels, local)

    if 'Sheet3' in wb.sheetnames:
        _write_sheet3(wb['Sheet3'], user, other)

    for ws in wb.worksheets:
        _configure_a4_print(ws)

    wb.save(fpath)

def generate_hrms_excel(fpath, user, diary, travel, hotels, local, other, attendance=None):
    # Keep this as is for system interoperability
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "HRMS"
    ws.append(["Staff No", user.staff_no])
    ws.append(["Name", user.name])
    wb.save(fpath)
