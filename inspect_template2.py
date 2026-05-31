"""Deeper dive: Sheet2 column structure, all data rows, and full column range."""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

PATH = r"C:\Users\nikhi\Desktop\AuditReport\template.xlsx"
wb = load_workbook(PATH, data_only=False)

ws = wb["Sheet2"]

print("=" * 120)
print("SHEET2: ALL NON-EMPTY COLUMNS (up to last used column)")
print("=" * 120)

# Find the last actual column with content
last_col = 0
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.value is not None:
            last_col = max(last_col, cell.column)

print(f"Last column with content: {last_col} ({get_column_letter(last_col)})")
print(f"Max row: {ws.max_row}")
print()

# Print all headers (rows 1-12) to understand columns
print("--- HEADER ROWS (1-12) ---")
for row_idx in range(1, 13):
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            print(f"  {col_letter}{row_idx}: {repr(cell.value)}")
print()

# Print ALL data rows (13-77) -- compact format
print("--- DATA ROWS (13-77) ---")
for row_idx in range(13, 78):
    vals = []
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            vals.append(f"{col_letter}{row_idx}={repr(cell.value)}")
    if vals:
        print(f"  Row {row_idx}: {'; '.join(vals)}")
print()

# Print rows 78-84 (summary area)
print("--- SUMMARY ROWS (78-84) ---")
for row_idx in range(78, 85):
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            print(f"  {col_letter}{row_idx}: {repr(cell.value)}")
print()

# Also check if there are any cells in columns beyond S in rows 1-84
print("--- ANY CONTENT BEYOND COLUMN S (cols T+) ---")
found = False
for row_idx in range(1, 85):
    for col_idx in range(20, ws.max_column + 1):  # T = 20
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            print(f"  {col_letter}{row_idx}: {repr(cell.value)}")
            found = True
if not found:
    print("  (nothing beyond column S)")

wb.close()
print("\nDone.")
