"""Check columns K through Q in Sheet2 data rows for which rows have values."""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\nikhi\Desktop\AuditReport\template.xlsx"
wb = load_workbook(PATH, data_only=False)
ws = wb["Sheet2"]

cols_to_check = ['K', 'L', 'M', 'N', 'O', 'P', 'Q']
col_indices = {c: ord(c) - ord('A') + 1 for c in cols_to_check}

print("=" * 120)
print("SHEET2: DATA IN COLUMNS K-R (rows 13-84)")
print("=" * 120)
for row_idx in range(13, 85):
    vals = []
    for col_letter, col_idx in col_indices.items():
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            vals.append(f"{col_letter}{row_idx}={repr(cell.value)}")
    # Also check R
    r_cell = ws.cell(row=row_idx, column=18)
    if r_cell.value is not None:
        vals.append(f"R{row_idx}={repr(r_cell.value)}")
    if vals:
        print(f"  Row {row_idx}: {'; '.join(vals)}")

print()
print("=" * 120)
print("SHEET2: CHECK IF ROWS 13-30 have values in columns L, N, O, P, Q")
print("=" * 120)
for row_idx in range(13, 31):
    for col_letter in ['L', 'N', 'O', 'P', 'Q']:
        cell = ws.cell(row=row_idx, column=ord(col_letter) - ord('A') + 1)
        if cell.value is not None:
            print(f"  {col_letter}{row_idx}: {repr(cell.value)}")

print()
print("=" * 120)
print("SHEET2: ALL EMPTY DATA ROWS (no content at all)")
print("=" * 120)
for row_idx in range(13, 78):
    has_content = False
    for col_idx in range(1, 21):
        if ws.cell(row=row_idx, column=col_idx).value is not None:
            has_content = True
            break
    if not has_content:
        print(f"  Row {row_idx} is completely empty")

print()
print("=" * 120)
print("SHEET2: Check rows 75-77 (between last data and grand total)")
print("=" * 120)
for row_idx in range(75, 78):
    has_content = False
    for col_idx in range(1, 21):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            col_letter = get_column_letter(col_idx)
            print(f"  {col_letter}{row_idx}: {repr(cell.value)}")
            has_content = True
    if not has_content:
        print(f"  Row {row_idx}: (empty)")

wb.close()
print("\nDone.")
